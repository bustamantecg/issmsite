from django.db.models import Q
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, Http404
from django.core.paginator import Paginator
from django.urls import reverse
import openpyxl
from openpyxl.styles import Alignment, Font
from django.views.generic import TemplateView


# Create your views here.
from secretaria.forms import AlumnoInscripCarreraForm, PersonaForm, CarreraForm
from secretaria.models import Carrera, Materia, Alumno, Persona, Docente, Empleado


#--------------Views para Secretaria -----------------------------------------------------#

@login_required
def secretaria(reeques):
    return render(reeques, 'secretaria/secretaria.html')


#--------------Views para Carrreras---------------------------------------------------------#
def carrera_listado(request):
    busqueda = request.POST.get('buscar')
    carreras = Carrera.objects.all().order_by('nombre')
    cantidad = len(carreras)
    if busqueda:
        carreras = Carrera.objects.filter(
            Q(nombre__icontains=busqueda) |
            Q(legal__icontains=busqueda) |
            Q(cuatrimestre__icontains=busqueda) |
            Q(tipo__icontains=busqueda) |
            Q(duracion__icontains=busqueda)
        ).distinct().order_by('nombre')
    return render(request, 'secretaria/carrera_listado.html', {'carreras': carreras, 'cantidad': cantidad})


def carrera_detalle(request, id):
    carrera = Carrera.objects.get(pk=id)
    materias = Materia.objects.filter(carrera_id=id).order_by('cuatrimestre', 'nombre')
    canti_materias = Materia.objects.filter(carrera_id=id).count()
    return render(request, 'secretaria/carrera_detalle.html', {'carrera': carrera, 'materias': materias, 'canti_materias': canti_materias})


def carrera_agregar(request):
    if request.method == 'POST':
        carrera_form = CarreraForm(request.POST)
        if carrera_form.is_valid():
            carrera_form.save()
            return redirect('carrera_listado')
    else:
        carrera_form = CarreraForm()
    return render(request, 'secretaria/carrera_agregar.html', {'carrera_form':carrera_form})


#--------------Views para Alumnos ---------------------------------------------------------

def alumno_listado(request):
    busqueda = request.GET.get('buscar')
    personas = Alumno.objects.all().order_by('persona__apellido', 'persona__nombre')
    cantidad = len(personas)
    paginate_by = 3
    if busqueda:
        personas = Alumno.objects.filter(
            Q(persona__dni__icontains=busqueda) |
            Q(persona__apellido__icontains=busqueda) |
            Q(persona__nombre__icontains=busqueda) |
            Q(persona__domicilio__icontains=busqueda) |
            Q(legajo__icontains=busqueda) |
            Q(persona__email__icontains=busqueda)
        ).distinct().order_by('persona__apellido')

        paginacion = Paginator(personas, 4)
        page = request.GET.get('page')
        personas = paginacion.get_page(page)

    return render(request, 'secretaria/alumno_listado.html', {'personas': personas, 'cantidad': cantidad})


def alumno_inscribir_carrera(request):
    if request.method == 'POST':
        formaAlumnoInscripCarrera = AlumnoInscripCarreraForm(request.POST)
        if formaAlumnoInscripCarrera.is_valid():

            persona = request.POST.get('persona')
            carrera = request.POST.get('carrera')
            existe = Alumno.objects.filter(persona_id= persona, carrera_id=carrera)

            if existe:
                return redirect(reverse(alumno_inscribir_carrera) + "?existe")
            

            formaAlumnoInscripCarrera.save()            
            return redirect(reverse(alumno_inscribir_carrera) + "?ok")
    else:
        formaAlumnoInscripCarrera = AlumnoInscripCarreraForm()
    return render(request, 'secretaria/alumno_inscribir_carrera.html',{'formaAlumnoInscripCarrera': formaAlumnoInscripCarrera})


def alumno_consulta(request):
    return render(request, 'secretaria/alumno_consulta.html')


class alumno_import_excel(TemplateView):
    def get(self, request, *args, **kwargs):
        alumnos= Alumno.objects.all().order_by('persona__apellido', 'persona__nombre')
        libro_excel = openpyxl.Workbook()
        hoja_excel= libro_excel.active
        #wb['B2']= 'Listado de Alumnos'
        #hoja_excel.cell(row= 1, column=2).value='Listado de Alumnos'
        ft = Font(bold=True, color='000000FF')
        titulo = 'Listado de Alumnos' # Título deseado
        celda_titulo = hoja_excel.cell(row=1, column=2, value=titulo)
        celda_titulo.alignment = Alignment(horizontal='center')  # Alineación centrada
        celda_titulo.font = ft

    # Combinar celdas para el título
        hoja_excel.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5) 

        #wb.merge_cells('B2:E2')

        enca= hoja_excel.cell(row= 3, column=2, value='DNI')       #   wb['B4']='DNI'
        enca.font= ft
        enca= hoja_excel.cell(row= 3, column=3, value='Apellido')  #   wb['C4']='Apellido'
        enca.font= ft
        enca= hoja_excel.cell(row= 3, column=4, value='Nombres' )  #   wb['D4']= 'Nombres' 
        enca.font= ft
        enca= hoja_excel.cell(row= 3, column=5, value='email')     #   wb['E4']= 'email' 
        enca.font= ft

        cont=4
        for alumno in alumnos:
            hoja_excel.cell(row= cont, column=2).value=alumno.persona.dni
            hoja_excel.cell(row= cont, column=3).value=alumno.persona.apellido
            hoja_excel.cell(row= cont, column=4).value=alumno.persona.nombre
            hoja_excel.cell(row= cont, column=5).value=alumno.persona.email
            cont+=1

        nombre_archivo= 'Listado_Alumnos.xlsx'
        #response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        response = HttpResponse(content_type= 'application/ms-excel')
        content= 'attachment; filename= {0}'.format(nombre_archivo)
        response['Content-Disposition']= content
        libro_excel.save(response)
        return response


def buscar_alumno_dni(request):
    if request.GET['dni']:
        dnibuscar = request.GET['dni']
        persona = Persona.objects.get(dni=dnibuscar)
        alumno = Alumno.objects.filter(persona=persona)

        carreras = Carrera.objects.filter(nombre__contains=alumno)
                #query = Alumno.objects.filter(persona__dni=dni).filter(carrera_id=Carrera.id)
        #return render(request, 'secretaria/alumno_resul_dni.html', {'persona': persona, 'query_dni': dni, 'alumno': alumno})
        return render(request, 'secretaria/alumno_resul_dni.html', {'persona': persona, 'query_dni': dnibuscar, 'alumno': alumno, 'carreras':carreras})
    else:
        mensaje = 'No ingreso D.N.I. del alumno/a a consultar'

    return HttpResponse(mensaje)


#--------------Views para Empleados de la Institutcion -----------------------------------------------#
def empleado_listado(request):
    busqueda = request.POST.get('buscar')
    empleados = Empleado.objects.all().order_by('persona__apellido')
    cantidad = len(empleados)
    encontrados = cantidad
    if busqueda:
        empleados = Empleado.objects.filter(
            Q(persona__dni__icontains=busqueda) |
            Q(persona__apellido__icontains=busqueda) |
            Q(persona__nombre__icontains=busqueda) |
            Q(cargo__icontains=busqueda) |
            Q(persona__email__icontains=busqueda)
        ).distinct().order_by('persona__apellido', 'persona__nombre')
        encontrados = len(empleados)
    return render(request, 'secretaria/empleado_listado.html', {'empleados': empleados, 'cantidad': cantidad, 'encontrados':encontrados})


#--------------Views para Personas---------------------------------------------------------#
def personas_listado(request):
    busqueda = request.POST.get('buscar')
    personas = Persona.objects.all().order_by('apellido', 'nombre')
    cantidad = len(personas)
    encontrados = cantidad
    if busqueda:
        personas = Persona.objects.filter(
            Q(dni__icontains=busqueda) |
            Q(apellido__icontains=busqueda) |
            Q(nombre__icontains=busqueda) |
            Q(dni__icontains=busqueda) |
            Q(email__icontains=busqueda)
        ).distinct().order_by('apellido', 'nombre')
        encontrados = len(personas)
    return render(request, 'secretaria/persona_listado.html', {'personas': personas, 'cantidad': cantidad, 'encontrados':encontrados})


def persona_agregar(request):
    titulo = 'Persona - Agregar'
    if request.method == 'POST':
        persona_form = PersonaForm(request.POST, request.FILES)
        if persona_form.is_valid():
            persona_form.save()
            return redirect('personas_listado')
    else:
        persona_form = PersonaForm()
    return render(request, 'secretaria/persona_agregar.html', {'persona_form':persona_form, 'titulo':titulo})


def persona_editar(request, id):
    persona = get_object_or_404(Persona, pk=id)
    if request.method == 'POST':
        formaPersona = PersonaForm(request.POST, instance=persona)
        if formaPersona.is_valid():
            formaPersona.save()
            return redirect('personas_listado')
    else:
        formaPersona = PersonaForm(instance=persona)
    return render(request, 'secretaria/persona_editar.html', {'formaPersona': formaPersona})


def persona_detalle(request, id):
    persona = get_object_or_404(Persona, pk=id)
    docente = Docente.objects.filter(persona_id=persona.id)
    print(docente)
    if not docente:

        docente = ''
    return render(request, 'secretaria/persona_detalle.html',{'persona': persona, 'docente': docente})